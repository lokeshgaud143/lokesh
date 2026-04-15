
import json
from openpyxl import load_workbook
from pathlib import Path

base = Path(__file__).resolve().parent
xlsx = base / "lokesh_website_manager.xlsx"
output = base / "data" / "site-data.json"

wb = load_workbook(xlsx, data_only=True)

def rows_to_dict(sheet_name):
    ws = wb[sheet_name]
    data = {}
    for key, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if key:
            data[str(key)] = value if value is not None else ""
    return data

site = rows_to_dict("Site")
theme = rows_to_dict("Theme")
seo = rows_to_dict("SEO")
about = rows_to_dict("About")
contact = rows_to_dict("Contact")

stats = []
ws = wb["Stats"]
for row in ws.iter_rows(min_row=2, values_only=True):
    label, value = row[:2]
    if label and value:
        stats.append({"label": str(label), "value": str(value)})

services = []
ws = wb["Services"]
for row in ws.iter_rows(min_row=2, values_only=True):
    title, description, icon, badge, enabled = row[:5]
    if str(enabled).strip().lower() in {"yes","true","1"} and title:
        services.append({
            "title": str(title),
            "description": str(description or ""),
            "icon": str(icon or "spark"),
            "badge": str(badge or "Service")
        })

projects = []
ws = wb["Projects"]
for row in ws.iter_rows(min_row=2, values_only=True):
    name, category, summary, link, tag, enabled = row[:6]
    if str(enabled).strip().lower() in {"yes","true","1"} and name:
        projects.append({
            "name": str(name),
            "category": str(category or ""),
            "summary": str(summary or ""),
            "link": str(link or "#"),
            "tag": str(tag or "Project")
        })

process = []
ws = wb["Process"]
for row in ws.iter_rows(min_row=2, values_only=True):
    step, title, text, enabled = row[:4]
    if str(enabled).strip().lower() in {"yes","true","1"} and step:
        process.append({
            "step": str(step),
            "title": str(title or ""),
            "text": str(text or "")
        })

socials = []
ws = wb["Socials"]
for row in ws.iter_rows(min_row=2, values_only=True):
    label, url, enabled = row[:3]
    if str(enabled).strip().lower() in {"yes","true","1"} and label and url:
        socials.append({"label": str(label), "url": str(url)})

contact_services = [item["title"] for item in services]

data = {
    "site": site,
    "theme": theme,
    "seo": seo,
    "about": {
        "heading": about.get("heading", ""),
        "body": about.get("body", ""),
        "stats": stats,
    },
    "services": services,
    "projects": projects,
    "process": process,
    "contact": {
        "heading": contact.get("heading", ""),
        "subheading": contact.get("subheading", ""),
        "services_list": contact_services,
        "form_endpoint": contact.get("form_endpoint", ""),
        "success_redirect": contact.get("success_redirect", "index.html#home"),
    },
    "socials": socials,
}
output.parent.mkdir(parents=True, exist_ok=True)
output.write_text(json.dumps(data, indent=2), encoding="utf-8")
print(f"Exported: {output}")
